from django.contrib import admin

import unicodecsv
import openpyxl
from functools import wraps
from collections import OrderedDict
from django.http import HttpResponse
from family.models import Person
from family.models import Family
from django.utils.translation import ugettext_lazy as _
from django.contrib.admin import SimpleListFilter
from singledispatch import singledispatch  # pip install singledispatch

from django.contrib.admin import SimpleListFilter
from django.utils.encoding import force_text
from django.utils.translation import ugettext as _
from django.core.urlresolvers import reverse
from django.utils.safestring import mark_safe
from django.db.models.query import QuerySet
from django.db.models.fields.files import FieldFile
from django.db.models.query import QuerySet
from django.db.models.fields.files import FieldFile
from openpyxl.cell import get_column_letter

def prep_field(request, obj, field, manyToManySep=';'):
    """ Returns the field as a unicode string. If the field is a callable, it
    attempts to call it first, without arguments.
    """
    if '__' in field:
        bits = field.split('__')
        field = bits.pop()
 
        for bit in bits:
            obj = getattr(obj, bit, None)
 
            if obj is None:
                return ""
 
    attr = getattr(obj, field)
    
    if isinstance(attr, (FieldFile,) ):
        attr = request.build_absolute_uri(attr.url)
        
    output = attr() if callable(attr) else attr
    
    if isinstance(output, (list, tuple, QuerySet)):
        output = manyToManySep.join([str(item) for item in output])
    return unicode(output).encode('utf-8') if output else ""
 
def export_csv_action(description="Export as CSV", fields=None, exclude=None, header=True,
                      manyToManySep=';'):
    """ This function returns an export csv action. """
    def export_as_csv(modeladmin, request, queryset):
        """ Generic csv export admin action.
        Based on http://djangosnippets.org/snippets/2712/
        """
        opts = modeladmin.model._meta
        field_names = [field.name for field in opts.fields]
        labels = []
 
        if exclude:
            field_names = [f for f in field_names if f not in exclude]
 
        elif fields:
            field_names = [field for field, _ in fields]
            labels = [label for _, label in fields]
 
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=%s.csv' % (
                unicode(opts).replace('.', '_')
            )
 
        writer = unicodecsv.writer(response)
 
        if header:
            writer.writerow(labels if labels else field_names)
 
        for obj in queryset:
            writer.writerow([prep_field(request, obj, field, manyToManySep) for field in field_names])
        return response
    export_as_csv.short_description = description
    export_as_csv.acts_on_all = True
    return export_as_csv 

def export_excel_action(description="Export as Excel", fields=None, exclude=None, header=True,
                      manyToManySep=';'):
    """ This function returns an export csv action. """
    def export_as_excel(modeladmin, request, queryset):
        """ Generic csv export admin action.
        Based on http://djangosnippets.org/snippets/2712/
        """
        opts = modeladmin.model._meta
        field_names = [field.name for field in opts.fields]
        labels = []
 
        if exclude:
            field_names = [f for f in field_names if f not in exclude]
 
        elif fields:
            field_names = [field for field, _ in fields]
            labels = [label for _, label in fields]

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name = "family_person.xlsx";
        response['Content-Disposition'] = 'attachment; filename='+file_name;
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "contact" 
    
        row_num = 0
        if header:
            col_num = 0
            for headrname in labels:
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = headrname
                col_num=col_num+1
        row_num=row_num+1

        for obj in queryset:
            col_num = 0;
            for field in field_names:
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = prep_field(request, obj, field)
                col_num=col_num+1
            row_num=row_num+1
    
        wb.save(response)
        return response

    export_as_excel.short_description = description
    export_as_excel.acts_on_all = True
    return export_as_excel 

def add_link_field(target_model = None, field = '', app='', field_name='link',
                   link_text=unicode):
    def add_link(cls):
        reverse_name = target_model or cls.model.__name__.lower()
        def link(self, instance):
            app_name = app or instance._meta.app_label
            reverse_path = "admin:%s_%s_change" % (app_name, reverse_name)
            link_obj = getattr(instance, field, None) or instance
            url = reverse(reverse_path, args = (link_obj.id,))
            return mark_safe("<a href='%s'>%s</a>" % (url, link_text(link_obj)))
        link.allow_tags = True
        link.short_description = reverse_name 
        setattr(cls, field_name, link)
        cls.readonly_fields = list(getattr(cls, 'readonly_fields', [])) + \
            [field_name]
        return cls
    return add_link
    
class PersonInline(admin.StackedInline):
    model = Person
    extra = 0

class FamilyAdmin(admin.ModelAdmin):
    list_display = ('id','address','city','state','zip','status', 'home1','home2', 'homefax')
    list_filter = ['status','city','state']
    search_fields = ['id','address','city','home1']
    inlines = [PersonInline]
        
@add_link_field('family','family',field_name='link2')
class PersonAdmin(admin.ModelAdmin):
    list_display = ('id','last','first','middle','chinese','mccc_dir','sex','role','link2','get_family_phone','email','cphone','wphone','worship','fellowship','fellowship2','baptized','bapday','category','birthday','member','memday','get_family_id','get_family_status')
    list_filter = ['fellowship','fellowship2','worship','baptized','member', 'family__status', 'category',]
    search_fields = ['last','first','chinese','email','comment', "family__address", "family__city", "family__state", "family__zip","family__home1","family__home2","family__homefax",]
    raw_id_fields = ("family",)
    
    def get_family_status(self, obj):
        return obj.family.status
    get_family_status.short_description = 'Family Status'
    get_family_status.admin_order_field = 'family__status'

    def get_family_id(self, obj):
        return obj.family.id
    get_family_id.short_description = 'Family Id'
    get_family_id.admin_order_field = 'family__id'

    def get_family_address(self, obj):
        list1 = [obj.family.address, obj.family.city, obj.family.state, obj.family.zip]
        faddress =[x for x in list1 if x is not None]
        return u','.join(faddress).encode('utf-8').strip()
        
    get_family_address.short_description = 'Family Address'
    get_family_address.admin_order_field = 'family__address'

    def get_family_phone(self, obj):
        list1 = [obj.family.home1, obj.family.home2, obj.family.homefax]
        fphones =[x for x in list1 if x is not None]
        return u' '.join(fphones).encode('utf-8').strip()
    get_family_phone.short_description = 'Family Phone'
    get_family_phone.admin_order_field = 'family__home1s'
    actions = [
        export_excel_action("Download selected objects as Excels file",
            fields=[
                ('last', 'Last Name'),
                ('first', 'First Name'),
                ('chinese','Chinese Name'),
                ('sex','Gender'),
                ('family__home1', 'Home Phone'),
                ('family__address', 'Address'),
                ('family__city', 'City'),
                ('family__state', 'State'),
                ('family__zip', 'Zip'),
                ('email', 'Email'),
                ('cphone', 'Cell Phone'),
                ('worship', 'Worship'),
                ('fellowship', 'Fellowship'),
                ('fellowship2', 'Fellowship2'),
                ('baptized', 'Baptized'),
                ('bapday', 'Bapday'),
                ('category', 'Category'),
                ('birthday', 'Birthday'),
            ],
            header=True,
            manyToManySep=';'
        ),
    ]
    

admin.site.register(Family,FamilyAdmin)
admin.site.register(Person,PersonAdmin)
