import unicodecsv
from functools import wraps
from collections import OrderedDict
from django.http import HttpResponse
from singledispatch import singledispatch  # pip install singledispatch
from django.contrib import admin
from django.contrib.admin import AdminSite
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.shortcuts import redirect
from children.models import CmMaster
from family.models import Person
import collections
#import autocomplete_light

def prep_field(obj, field):
    """
    (for download_as_csv action)
    Returns the field as a unicode string. If the field is a callable, it
    attempts to call it first, without arguments.
    """
    if '__' in field:
        bits = field.split('__')
        field = bits.pop()

        for bit in bits:
            obj = getattr(obj, bit, None)

            if obj is None:
                return ""

    attr = getattr(obj, field)
    output = attr() if isinstance(attr, collections.Callable) else attr
    return str(output).encode('utf-8') if output is not None else ""

@singledispatch
def download_as_excel(modeladmin, request, queryset):
    """
    Generic exel export admin action.

    Example:

        class ExampleModelAdmin(admin.ModelAdmin):
            raw_id_fields = ('field1',)
            list_display = ('field1', 'field2', 'field3',)
            actions = [download_as_csv,]
            download_as_csv_fields = [
                'field1',
                ('foreign_key1__foreign_key2__name', 'label2'),
                ('field3', 'label3'),
            ],
            download_as_csv_header = True
    """
    fields = getattr(modeladmin, 'download_as_excel_fields', None)
    exclude = getattr(modeladmin, 'download_as_excel_exclude', None)
    header = getattr(modeladmin, 'download_as_excel_header', True)
    verbose_names = getattr(modeladmin, 'download_as_excel_verbose_names', True)

    opts = modeladmin.model._meta

    def fname(field):
        if verbose_names:
            return str(field.verbose_name).capitalize()
        else:
            return field.name

    # field_names is a map of {field lookup path: field label}
    if exclude:
        field_names = OrderedDict(
            (f.name, fname(f)) for f in opts.fields if f not in exclude
        )
    elif fields:
        field_names = OrderedDict()
        for spec in fields:
            if isinstance(spec, (list, tuple)):
                field_names[spec[0]] = spec[1]
            else:
                try:
                    f, _, _, _ = opts.get_field_by_name(spec)
                except FieldDoesNotExist:
                    field_names[spec] = spec
                else:
                    field_names[spec] = fname(f)
    else:
        field_names = OrderedDict(
            (f.name, fname(f)) for f in opts.fields
        )

    # response = HttpResponse(content_type='text/csv')
    # response['Content-Disposition'] = 'attachment; filename=%s.csv' % (
    #         unicode(opts).replace('.', '_')
    #     )

    import openpyxl
    from openpyxl.cell import get_column_letter

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = "CM Master.xlsx";
    response['Content-Disposition'] = 'attachment; filename='+file_name;
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "CM Master" 

    row_num = 0
    col_num = 0;

    # writer = csv.writer(response)
    # writer = unicodecsv.writer(response, encoding='utf-8')
    
    if header:
        col_num = 0;
        for headrname in list(field_names.values()):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = headrname
            col_num=col_num+1
    row_num=row_num+1
    for obj in queryset:
        # writer.writerow([prep_field(obj, field) for field in field_names.keys()])
        col_num = 0;
        for field in list(field_names.keys()):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = prep_field(obj, field)
            col_num=col_num+1
        row_num=row_num+1

    wb.save(response)
    return response

download_as_excel.short_description = "Download selected objects as Excel file"

@download_as_excel.register(str)
def _(description):
    """
    (overridden dispatcher)
    Factory function for making a action with custom description.

    Example:

        class ExampleModelAdmin(admin.ModelAdmin):
            raw_id_fields = ('field1',)
            list_display = ('field1', 'field2', 'field3',)
            actions = [download_as_csv("Export Special Report"),]
            download_as_csv_fields = [
                'field1',
                ('foreign_key1__foreign_key2__name', 'label2'),
                ('field3', 'label3'),
            ],
            download_as_csv_header = True
    """
    @wraps(download_as_excel)
    def wrapped_action(modeladmin, request, queryset):
        return download_as_excel(modeladmin, request, queryset)
    wrapped_action.short_description = description
    return wrapped_action

def promote(x):
    return {
        'N': 'PK',
        'PK': 'K',
        'K': '1',
        '1': '2',
        '2': '3',
        '3': '4',
        '4': '5',
        '5': '6',
        '6': '7',
        '7': '8',
        '8': '9',
        '9': '10',
        '10': '11',
        '11': '12',
    }[x]


def grade_promotion(modeladmin, request, queryset):
    for child in queryset:
        child.ssgrade=promote(child.ssgrade);
        child.save()
 
grade_promotion.short_description = "Grade Promotion"

class ChildrenSite(AdminSite):
    site_header = 'Children'
    
    def login(self, request, extra_context=None):
        return redirect('%s?next=%s' % (reverse('home'), request.REQUEST.get('next', '')))

class CmMasterAdmin(admin.ModelAdmin):
    list_display = ['first_last','ssgrade','ssactive','fname','lname','chinese_name','gender','dob',
    'allergies_medical_conditions_medications','fathers_english_name','fathers_chinese_name_if_available','mothers_english_name','mother_chinese_name_if_available',
    'email','street','city','state','zip','home','fathers_office','fathers_cell','mothers_office','mothers_cell','alternate_contact_name','alt_contact_main_phone',
    'altcont','mccc','group','assign','christianfather','christianmother','remarks','felly','vbs_2010','vbs_2011','vbs_2012','vbs_2013','vbs_2014','vbs_2015','vbs_2016',]
    search_fields = ['first_last','fname','lname','chinese_name','allergies_medical_conditions_medications','fathers_english_name','fathers_chinese_name_if_available','mothers_english_name','mother_chinese_name_if_available',
    'email','street','city','state','zip','home']
    list_filter = ['ssactive','ssgrade','vbs_2016','vbs_2015','vbs_2014','vbs_2013','vbs_2012','vbs_2011','vbs_2010']
#    form = autocomplete_light.modelform_factory(CmMaster,exclude=[])
    actions = [download_as_excel("Download selected objects as Excel file"),grade_promotion,]
    download_as_excel_fields=['first_last',
       'ssgrade',
       'ssactive',
       'fname',
       'lname',
       'chinese_name',
       'gender',
       'grade',
       'dob',
       'allergies_medical_conditions_medications',
       'fathers_english_name',
       'fathers_chinese_name_if_available',
       'mothers_english_name',
       'mother_chinese_name_if_available',
       'email',
       'street',
       'city',
       'state',
       'zip',
       'home',
       'fathers_office',
       'fathers_cell',
       'mothers_office',
       'mothers_cell',
       'alternate_contact_name',
       'alt_contact_main_phone',
       'altcont',
       'mccc',
       'group',
       'assign',
       'christianfather',
       'christianmother',
       'remarks',
       'felly',         
       'vbs_2010',
       'vbs_2011',
       'vbs_2012',
       'vbs_2013',
       'vbs_2014',
       'vbs_2015',
       'vbs_2016',
    ]
    download_as_excel_header = True

    
# for the person raw_id picker widget 
# The raw_id_fields widget shows a magnifying glass button next to the field which allows users to search for and select a value
class PersonPicker(admin.ModelAdmin):
    list_display = ('id','last','first','chinese','sex','email', 'cphone','role','birthday',)
    search_fields = ['last','first','chinese','email', 'cphone',]

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return True

    def has_delete_permission(self, request, obj=None):
        return False

children_site = ChildrenSite(name='children')
        
children_site.register(CmMaster,CmMasterAdmin)
children_site.register(Person, PersonPicker)
